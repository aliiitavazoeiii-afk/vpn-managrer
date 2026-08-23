from __future__ import annotations
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .models import User


def _text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _norm(v: Any) -> str:
    return " ".join(_text(v).lower().split())


def _phone(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip().replace(" ", "").replace("-", "")
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    if s.isdigit() and len(s) == 10 and s.startswith("9"):
        s = "0" + s
    return s or None


def _amount_toman(v: Any) -> int:
    if v in (None, ""):
        return 200_000
    try:
        n = int(float(str(v).replace(",", "").strip()))
    except Exception:
        return 200_000
    if 0 < n < 10_000:
        n *= 1000
    return max(n, 0)


def _due_date(v: Any) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _status(v: Any) -> str:
    s = _norm(v)
    return "paid" if s in {"done", "paid", "yes", "true", "1", "پرداخت شد", "پرداخت شده"} else "unpaid"


def import_users_xlsx(db: Session, raw: bytes) -> dict:
    if not raw:
        raise ValueError("فایل خالی است")
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    existing = {_norm(u.name): u for u in db.query(User).all() if _norm(u.name)}
    created = updated = skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue
        vals = list(row) + [None] * 5
        name = _text(vals[0])
        if not name:
            skipped += 1
            continue
        key = _norm(name)
        user = existing.get(key)
        if user is None:
            user = User(name=name)
            db.add(user)
            existing[key] = user
            created += 1
        else:
            updated += 1

        user.name = name
        user.due_date = _due_date(vals[1])
        user.fee_toman = _amount_toman(vals[2])
        user.phone = _phone(vals[3])
        user.payment_status = _status(vals[4])

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "total": created + updated}
